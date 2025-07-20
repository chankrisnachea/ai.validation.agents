import json
from openpyxl import load_workbook
import re

# Load workbook (formulas visible)
wb = load_workbook("TCD_Baseline.xlsx", data_only=False)
ws = wb.active

# Extract headers
headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

data = []

# Regex to extract second argument (display text) from HYPERLINK
hyperlink_pattern = re.compile(r'HYPERLINK\((?:"|\')?.*?(?:"|\')?,\s*(?:"|\')(.*?)(?:"|\')\)', re.IGNORECASE)

for row in ws.iter_rows(min_row=2):
    row_dict = {}

    for header, cell in zip(headers, row):
        value = cell.value

        # Normalize and extract display text if it's a HYPERLINK formula
        if isinstance(value, str) and "HYPERLINK" in value.upper():
            normalized = value.lstrip('=').lstrip('=')  # Remove any leading ==
            match = hyperlink_pattern.search(normalized)
            if match:
                value = match.group(1)

        row_dict[header] = value

    data.append(row_dict)

# Write JSON
with open("tcd_baseline.json", "w") as f:
    json.dump(data, f, indent=4)

print("Saved output.json")
